from openpyxl import load_workbook
import os
import json

path = os.getcwd()
path += "\\xlsx"
os.chdir(path)
files = os.listdir()


lookup = {}
for singleFile in files:
    wb = load_workbook(singleFile)
    ws = wb.active
    rowNum = 3
    while ws.cell(row=rowNum,column=1).value :
        cityName = ws.cell(row=rowNum,column=4).value
        stateName = ws.cell(row=rowNum,column=2).value
        if stateName in lookup:
            if (cityName in lookup[stateName]) == False:
                lookup[stateName].append(cityName)
        else:
            lookup[stateName] = []
            lookup[stateName].append(cityName)
        rowNum += 1


jsonData = json.dumps(lookup)
os.chdir("..")
path = os.getcwd()
outFile = open("output.json","w")
outFile.write(jsonData)
outFile.close()
print("Done")